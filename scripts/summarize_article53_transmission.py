#!/usr/bin/env python3
"""Reanalyse species sharing and summarize published strain/FMT evidence."""

from __future__ import annotations

import argparse
import math
import re
import shutil
from collections import Counter
from pathlib import Path
from statistics import median

from lxml import etree
from openpyxl import load_workbook

from article41_44_utils import dump_json, write_tsv


THRESHOLDS = (0.0, 0.01, 0.1, 1.0)


def clean_text(root: etree._ElementTree) -> str:
    return " ".join(" ".join(root.getroot().itertext()).split())


def quantile(values: list[float], probability: float) -> float:
    ordered = sorted(values)
    if not ordered:
        raise ValueError("Cannot calculate quantile of an empty vector")
    position = (len(ordered) - 1) * probability
    lower = math.floor(position)
    upper = math.ceil(position)
    if lower == upper:
        return ordered[lower]
    fraction = position - lower
    return ordered[lower] * (1 - fraction) + ordered[upper] * fraction


def exact_sign_p(greater: int, less: int) -> float:
    n = greater + less
    tail = min(greater, less)
    return min(1.0, 2 * sum(math.comb(n, i) for i in range(tail + 1)) / (2**n))


def as_number(value: object) -> float | None:
    if value in (None, "NA", ""):
        return None
    return float(value)


def parse_filtered_millions(value: object) -> float | None:
    if value in (None, "NA", ""):
        return None
    match = re.search(r"\(([-0-9.]+)\s*M?\)", str(value))
    return float(match.group(1)) if match else None


def mother_design(work: Path, output: Path) -> list[dict[str, object]]:
    workbook = load_workbook(
        work / "input/sys001172080st8.xlsx", read_only=True, data_only=True
    )
    rows = list(workbook.active.iter_rows(values_only=True))[3:]
    result = []
    for row in rows:
        if row[0] is None:
            continue
        age_match = re.search(r"\((\d+)\s*mo\.\)", str(row[1]))
        time_match = re.match(r"(\d+)", str(row[1]))
        result.append(
            {
                "Pair": int(row[0]),
                "TimePoint": f"T{time_match.group(1)}",
                "InfantAgeMonths": int(age_match.group(1)),
                "InfantSex": row[2],
                "BreastFeedingFraction": row[3],
                "FormulaFeedingFraction": row[4],
                "AfterWeaning": row[5],
                "MotherFilteredReadsMillion": parse_filtered_millions(row[11]),
                "InfantFilteredReadsMillion": parse_filtered_millions(row[12]),
                "MilkFilteredReadsMillion": parse_filtered_millions(row[13]),
            }
        )
    if len(result) != 8 or len({row["Pair"] for row in result}) != 5:
        raise ValueError("Unexpected Asnicar Table S1 design")
    write_tsv(output / "mother-infant-design.tsv", result)
    return result


def mother_species(work: Path, output: Path) -> dict[str, object]:
    workbook = load_workbook(
        work / "input/sys001172080st9.xlsx", read_only=True, data_only=True
    )
    rows = list(workbook["all"].iter_rows(values_only=True))
    samples = [str(value) for value in rows[0][1:]]
    metadata = {
        str(rows[index][0]): list(rows[index][1:]) for index in range(1, 4)
    }
    if len(samples) != 24 or Counter(metadata["sample_type"]) != Counter(
        {"Infant": 8, "Mother": 8, "Milk": 8}
    ):
        raise ValueError("Unexpected Asnicar Table S2 sample coordinate")
    sample_meta = {}
    for index, sample in enumerate(samples):
        sample_meta[sample] = {
            "Sample": sample,
            "SampleType": metadata["sample_type"][index],
            "TimePoint": metadata["time_point"][index],
            "Pair": int(metadata["pair"][index]),
        }
    write_tsv(output / "mother-infant-sample-metadata.tsv", sample_meta.values())

    species: list[tuple[str, str, list[float]]] = []
    profile_rows: list[dict[str, object]] = []
    for row in rows[4:]:
        taxonomy = str(row[0] or "")
        if "|s__" not in taxonomy or "|t__" in taxonomy:
            continue
        leaf = taxonomy.split("|")[-1]
        if leaf.endswith("_unclassified"):
            continue
        name = leaf.removeprefix("s__")
        values = [float(value or 0) for value in row[1:]]
        species.append((name, taxonomy, values))
        for sample, value in zip(samples, values):
            profile_rows.append(
                {
                    "Species": name,
                    "Taxonomy": taxonomy,
                    "Sample": sample,
                    "RelativeAbundancePct": value,
                }
            )
    if len(species) != 247:
        raise ValueError(f"Expected 247 named species rows, observed {len(species)}")
    write_tsv(output / "mother-infant-species-profile.tsv.gz", profile_rows)

    infants = [s for s in samples if sample_meta[s]["SampleType"] == "Infant"]
    mothers = [s for s in samples if sample_meta[s]["SampleType"] == "Mother"]
    pairwise_rows: list[dict[str, object]] = []
    rank_rows: list[dict[str, object]] = []
    threshold_summaries: list[dict[str, object]] = []
    for threshold in THRESHOLDS:
        present = {
            sample: {
                name
                for name, _, values in species
                if values[samples.index(sample)] > threshold
            }
            for sample in samples
        }
        current = []
        for infant in infants:
            imeta = sample_meta[infant]
            eligible = [
                mother
                for mother in mothers
                if sample_meta[mother]["TimePoint"] == imeta["TimePoint"]
            ]
            infant_rows = []
            for mother in eligible:
                mmeta = sample_meta[mother]
                shared = len(present[infant] & present[mother])
                union = len(present[infant] | present[mother])
                jaccard = shared / union if union else 0.0
                entry = {
                    "ThresholdPctExclusive": threshold,
                    "ThresholdLabel": f">{threshold:g}%",
                    "Infant": infant,
                    "InfantPair": imeta["Pair"],
                    "TimePoint": imeta["TimePoint"],
                    "Mother": mother,
                    "MotherPair": mmeta["Pair"],
                    "MatchedPair": imeta["Pair"] == mmeta["Pair"],
                    "InfantDetectedSpecies": len(present[infant]),
                    "MotherDetectedSpecies": len(present[mother]),
                    "SharedSpecies": shared,
                    "JaccardSimilarity": round(jaccard, 8),
                }
                pairwise_rows.append(entry)
                current.append(entry)
                infant_rows.append(entry)
            matched = next(row for row in infant_rows if row["MatchedPair"])
            ordered = sorted(
                infant_rows,
                key=lambda row: (-float(row["JaccardSimilarity"]), str(row["Mother"])),
            )
            top_value = float(ordered[0]["JaccardSimilarity"])
            rank = 1 + sum(
                float(row["JaccardSimilarity"])
                > float(matched["JaccardSimilarity"])
                for row in infant_rows
            )
            tie_size = sum(
                math.isclose(float(row["JaccardSimilarity"]), top_value, abs_tol=1e-12)
                for row in infant_rows
            )
            rank_rows.append(
                {
                    "ThresholdPctExclusive": threshold,
                    "ThresholdLabel": f">{threshold:g}%",
                    "Infant": infant,
                    "InfantPair": imeta["Pair"],
                    "TimePoint": imeta["TimePoint"],
                    "CandidateMothers": len(infant_rows),
                    "UnrelatedNegativeControls": len(infant_rows) - 1,
                    "MatchedMother": matched["Mother"],
                    "MatchedJaccard": matched["JaccardSimilarity"],
                    "MatchedRank": rank,
                    "TopTieSize": tie_size,
                    "RankedFirstIncludingTies": math.isclose(
                        float(matched["JaccardSimilarity"]), top_value, abs_tol=1e-12
                    ),
                    "UniquelyRankedFirst": rank == 1 and tie_size == 1,
                }
            )
        matched_values = [row for row in current if row["MatchedPair"]]
        unrelated_values = [row for row in current if not row["MatchedPair"]]
        eligible_ranks = [
            row
            for row in rank_rows
            if row["ThresholdPctExclusive"] == threshold
            and int(row["UnrelatedNegativeControls"]) > 0
        ]
        threshold_summaries.append(
            {
                "ThresholdPctExclusive": threshold,
                "ThresholdLabel": f">{threshold:g}%",
                "MatchedComparisons": len(matched_values),
                "UnrelatedComparisons": len(unrelated_values),
                "MatchedMedianSharedSpecies": median(
                    int(row["SharedSpecies"]) for row in matched_values
                ),
                "UnrelatedMedianSharedSpecies": median(
                    int(row["SharedSpecies"]) for row in unrelated_values
                ),
                "MatchedMedianJaccard": round(
                    median(float(row["JaccardSimilarity"]) for row in matched_values), 8
                ),
                "UnrelatedMedianJaccard": round(
                    median(float(row["JaccardSimilarity"]) for row in unrelated_values), 8
                ),
                "InfantsWithNegativeControls": len(eligible_ranks),
                "MatchedRankedFirstIncludingTies": sum(
                    bool(row["RankedFirstIncludingTies"]) for row in eligible_ranks
                ),
                "MatchedUniquelyRankedFirst": sum(
                    bool(row["UniquelyRankedFirst"]) for row in eligible_ranks
                ),
            }
        )
    write_tsv(output / "mother-infant-pairwise-sharing.tsv", pairwise_rows)
    write_tsv(output / "mother-infant-rank-sensitivity.tsv", rank_rows)
    write_tsv(output / "mother-infant-sharing-summary.tsv", threshold_summaries)
    primary = next(row for row in threshold_summaries if row["ThresholdPctExclusive"] == 0.1)
    return {
        "samples": len(samples),
        "named_species": len(species),
        "primary": primary,
    }


def published_mother_strains(work: Path, output: Path) -> list[dict[str, object]]:
    root = etree.parse(str(work / "input/PMC5264247.xml"))
    text = clean_text(root)
    required = (
        "99.96% sequence identity",
        "99.87% intrapair similarity",
        "99.93% similarity",
        "independent acquisition of strains from a shared environmental source cannot be excluded",
        "vertical transmission from mother to infant for 14% of the species",
        "MUSCLE version v3.8.1551",
        "RAxML version 8.1.15",
        "-p 1234",
    )
    missing = [phrase for phrase in required if phrase not in text]
    if missing:
        raise ValueError(f"Pinned mother-infant XML lacks expected evidence: {missing}")
    rows = [
        {
            "Species": "Bifidobacterium bifidum",
            "Pair": 4,
            "TimePoint": "T2",
            "IntraPairSimilarityPct": 99.96,
            "IntraPairDivergencePct": 0.04,
            "ClosestOtherDivergencePct": 0.60,
            "ClosestOtherQualifier": "at least",
            "AverageOtherDivergencePct": "NA",
            "ReportedP": "4.7e-40",
            "SequenceEvidence": "StrainPhlAn species-specific marker SNVs",
            "IndependentGeneContentEvidence": "PanPhlAn Figure S5",
        },
        {
            "Species": "Coprococcus comes",
            "Pair": 5,
            "TimePoint": "reported shared time point",
            "IntraPairSimilarityPct": 99.87,
            "IntraPairDivergencePct": 0.13,
            "ClosestOtherDivergencePct": 1.60,
            "ClosestOtherQualifier": "reported closest",
            "AverageOtherDivergencePct": 1.61,
            "ReportedP": "1.9e-3",
            "SequenceEvidence": "StrainPhlAn species-specific marker SNVs",
            "IndependentGeneContentEvidence": "PanPhlAn Figure S5",
        },
        {
            "Species": "Ruminococcus bromii",
            "Pair": 5,
            "TimePoint": "reported shared time point",
            "IntraPairSimilarityPct": 99.93,
            "IntraPairDivergencePct": 0.07,
            "ClosestOtherDivergencePct": 1.53,
            "ClosestOtherQualifier": "reported closest",
            "AverageOtherDivergencePct": 2.63,
            "ReportedP": "4.9e-8",
            "SequenceEvidence": "StrainPhlAn species-specific marker SNVs",
            "IndependentGeneContentEvidence": "PanPhlAn Figure S5",
        },
    ]
    write_tsv(output / "published-mother-infant-strain-evidence.tsv", rows)
    write_tsv(
        output / "mother-infant-evidence-ledger.tsv",
        [
            {
                "Species": row["Species"],
                "MarkerSNVTree": True,
                "QuantitativeDivergenceReported": True,
                "GeneRepertoireCorroboration": True,
                "DirectionSupportedByTiming": True,
                "SharedEnvironmentExcluded": False,
            }
            for row in rows
        ]
        + [
            {
                "Species": species,
                "MarkerSNVTree": True,
                "QuantitativeDivergenceReported": False,
                "GeneRepertoireCorroboration": True,
                "DirectionSupportedByTiming": True,
                "SharedEnvironmentExcluded": False,
            }
            for species in (
                "Bifidobacterium adolescentis",
                "Bifidobacterium breve",
                "Bifidobacterium longum",
            )
        ],
    )
    return rows


def fmt_outputs(work: Path, output: Path) -> dict[str, object]:
    xml = etree.parse(str(work / "input/PMC8951724.xml"))
    text = clean_text(xml)
    required = (
        "default: ≥ 5 kb",
        "default: ≥ 99.9%",
        "relative allelic frequencies (default: ≥ 10%)",
        "MetaPhlAn2 v2.6.0",
        "Samtools v0.1.19",
        "kpileup v1.0",
        "db_v20, mpa_v20_m200",
    )
    missing = [phrase for phrase in required if phrase not in text]
    if missing:
        raise ValueError(f"Pinned SameStr XML lacks expected methods: {missing}")

    workbook = load_workbook(
        work / "input/40168_2022_1251_MOESM7_ESM.xlsx",
        read_only=True,
        data_only=True,
    )
    sample_sheet = workbook["S1. Sample Metadata"]
    sample_rows = list(sample_sheet.iter_rows(values_only=True))
    sample_header = sample_rows[0]
    samples = [dict(zip(sample_header, row)) for row in sample_rows[1:] if row[0] is not None]
    write_tsv(
        output / "fmt-sample-metadata.tsv.gz",
        [
            {
                "Study": row["Study"],
                "UniqueID": row["Unique_ID"],
                "Case": row["Case_Name"],
                "DaysSinceFMT": row["Days_Since_FMT"],
                "SampleType": row["Sample_Type"],
                "FMTSuccess": row["fmt_success"],
                "StudyType": row["Study_Type"],
            }
            for row in samples
        ],
    )

    case_sheet = workbook["S7. Case-wise Shared Taxa"]
    raw_cases = [
        row
        for row in case_sheet.iter_rows(min_row=3, values_only=True)
        if row[0] is not None
    ]
    cases = []
    long_rows = []
    for row in raw_cases:
        entry = {
            "Study": row[0],
            "Case": row[1],
            "FMTOutcome": row[2],
            "DaysSinceFMT": int(row[3]),
            "PrePostSharedStrains": as_number(row[4]),
            "DonorPostSharedStrains": as_number(row[5]),
            "PrePostSharedSpecies": as_number(row[6]),
            "DonorPostSharedSpecies": as_number(row[7]),
        }
        cases.append(entry)
        for resolution, pre_key, donor_key in (
            ("Strain", "PrePostSharedStrains", "DonorPostSharedStrains"),
            ("Species", "PrePostSharedSpecies", "DonorPostSharedSpecies"),
        ):
            for comparison, key in (
                ("Pre-FMT / post-FMT", pre_key),
                ("Donor / post-FMT", donor_key),
            ):
                if entry[key] is not None:
                    long_rows.append(
                        {
                            "Study": entry["Study"],
                            "Case": entry["Case"],
                            "FMTOutcome": entry["FMTOutcome"],
                            "DaysSinceFMT": entry["DaysSinceFMT"],
                            "Resolution": resolution,
                            "Comparison": comparison,
                            "SharedTaxa": entry[key],
                        }
                    )
    if len(cases) != 27:
        raise ValueError(f"Expected 27 FMT cases, observed {len(cases)}")
    write_tsv(output / "fmt-casewise-sharing.tsv", cases)
    write_tsv(output / "fmt-casewise-sharing-long.tsv", long_rows)

    paired_summary = []
    for resolution, pre_key, donor_key in (
        ("Strain", "PrePostSharedStrains", "DonorPostSharedStrains"),
        ("Species", "PrePostSharedSpecies", "DonorPostSharedSpecies"),
    ):
        paired = [
            (float(row[pre_key]), float(row[donor_key]))
            for row in cases
            if row[pre_key] is not None and row[donor_key] is not None
        ]
        pre = [left for left, _ in paired]
        donor = [right for _, right in paired]
        deltas = [right - left for left, right in paired]
        greater = sum(delta > 0 for delta in deltas)
        equal = sum(delta == 0 for delta in deltas)
        less = sum(delta < 0 for delta in deltas)
        paired_summary.append(
            {
                "Resolution": resolution,
                "CompleteCases": len(paired),
                "PrePostMedian": median(pre),
                "PrePostQ1": quantile(pre, 0.25),
                "PrePostQ3": quantile(pre, 0.75),
                "DonorPostMedian": median(donor),
                "DonorPostQ1": quantile(donor, 0.25),
                "DonorPostQ3": quantile(donor, 0.75),
                "MedianPairedDifference": median(deltas),
                "DonorGreater": greater,
                "Equal": equal,
                "DonorLess": less,
                "ExactTwoSidedSignP": round(exact_sign_p(greater, less), 12),
            }
        )
    write_tsv(output / "fmt-casewise-sharing-summary.tsv", paired_summary)

    classifier_sheet = workbook["S6. Logistic Regression"]
    classifier_rows = []
    for row in classifier_sheet.iter_rows(min_row=8, max_row=11, values_only=True):
        classifier_rows.extend(
            [
                {
                    "TaxonomicLevel": str(row[0]).title(),
                    "TestSet": "Control hold-out",
                    "AUROC": float(row[1]),
                    "AUPR": float(row[2]),
                },
                {
                    "TaxonomicLevel": str(row[0]).title(),
                    "TestSet": "rCDI / FMT",
                    "AUROC": float(row[3]),
                    "AUPR": float(row[4]),
                },
            ]
        )
    write_tsv(output / "relatedness-classifier-performance.tsv", classifier_rows)

    event_sheet = workbook["S8. Competing Strain Events"]
    event_iterator = event_sheet.iter_rows(values_only=True)
    event_header = [str(value) for value in next(event_iterator)]
    raw_events = [
        dict(zip(event_header, row)) for row in event_iterator if row[0] is not None
    ]
    event_rows = []
    for row in raw_events:
        event_rows.append(
            {
                "Study": row["Study"],
                "Case": row["Case_Name"],
                "DaysSinceFMT": row["Days_Since_FMT.post"],
                "FMTSuccess": row["fmt_success"],
                "Species": row["species"],
                "Source": row["source"],
                "RecipientRelativeAbundancePct": row["rel_abund.recipient"],
                "PostRelativeAbundancePct": row["rel_abund.post"],
                "DonorRelativeAbundancePct": row["rel_abund.donor"],
                "PrePostMVS": row["Pre-FMT/Post-FMT.mvs"],
                "DonorPostMVS": row["Donor/Post-FMT.mvs"],
                "PrePostOverlapBp": row["Pre-FMT/Post-FMT.overlap"],
                "DonorPostOverlapBp": row["Donor/Post-FMT.overlap"],
            }
        )
    if len(event_rows) != 408:
        raise ValueError(f"Expected 408 competing-strain events, observed {len(event_rows)}")
    write_tsv(output / "fmt-competing-strain-events.tsv.gz", event_rows)

    source_rows = []
    for success, label in ((True, "Resolved"), (False, "Failed")):
        subset = [row for row in event_rows if bool(row["FMTSuccess"]) == success]
        counts = Counter(str(row["Source"]) for row in subset)
        for source in ("donor", "self", "both", "unique"):
            source_rows.append(
                {
                    "FMTOutcome": label,
                    "Source": source.title(),
                    "Events": counts[source],
                    "TotalEvents": len(subset),
                    "EventFraction": round(counts[source] / len(subset), 8),
                    "Cases": len({row["Case"] for row in subset}),
                }
            )
    write_tsv(output / "fmt-source-event-summary.tsv", source_rows)

    pure_category_audit = []
    for source in ("donor", "self", "unique"):
        subset = [row for row in event_rows if row["Source"] == source]
        for row in subset:
            donor_call = (
                float(row["DonorPostMVS"]) >= 0.999
                and float(row["DonorPostOverlapBp"]) >= 5000
            )
            self_call = (
                float(row["PrePostMVS"]) >= 0.999
                and float(row["PrePostOverlapBp"]) >= 5000
            )
            expected = (
                donor_call and not self_call
                if source == "donor"
                else self_call and not donor_call
                if source == "self"
                else not donor_call and not self_call
            )
            pure_category_audit.append(
                {
                    "Source": source,
                    "Case": row["Case"],
                    "Species": row["Species"],
                    "DonorGate": donor_call,
                    "SelfGate": self_call,
                    "ExpectedPureCategory": expected,
                }
            )
    if not all(row["ExpectedPureCategory"] for row in pure_category_audit):
        raise ValueError("Pure SameStr source categories do not reproduce the hard gates")
    write_tsv(output / "fmt-pure-source-gate-audit.tsv.gz", pure_category_audit)

    return {
        "workbook_samples": len(samples),
        "rcdi_samples": sum(row["Study_Type"] == "rCDI" for row in samples),
        "control_samples": sum(row["Study_Type"] == "Control" for row in samples),
        "cases": len(cases),
        "complete_paired_cases": 25,
        "events": len(event_rows),
        "source_counts": dict(Counter(str(row["Source"]) for row in event_rows)),
        "paired_summary": {row["Resolution"]: row for row in paired_summary},
    }


def evidence_ladder(output: Path) -> None:
    write_tsv(
        output / "transmission-evidence-ladder.tsv",
        [
            {
                "Level": 1,
                "Evidence": "Same species",
                "MinimumData": "taxonomic profiles",
                "ClaimAllowed": "co-occurrence only",
                "MainFailureMode": "ubiquitous species and unrelated strains",
            },
            {
                "Level": 2,
                "Evidence": "Same strain",
                "MinimumData": "SNV similarity plus callable overlap/coverage",
                "ClaimAllowed": "strain sharing",
                "MainFailureMode": "threshold, reference and mixed-strain bias",
            },
            {
                "Level": 3,
                "Evidence": "Temporal direction",
                "MinimumData": "source before recipient and longitudinal follow-up",
                "ClaimAllowed": "transmission-compatible direction",
                "MainFailureMode": "unsampled source or transient passage",
            },
            {
                "Level": 4,
                "Evidence": "Negative-source controls",
                "MinimumData": "unrelated pairs plus household/environment/food controls",
                "ClaimAllowed": "direct-source inference strengthened",
                "MainFailureMode": "residual shared-environment confounding",
            },
            {
                "Level": 5,
                "Evidence": "Colonization and corroboration",
                "MinimumData": "persistence plus independent genome/gene-content evidence",
                "ClaimAllowed": "engraftment or durable transmission",
                "MainFailureMode": "replacement, detection limit and functional ambiguity",
            },
        ],
    )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--project-root", type=Path, required=True)
    parser.add_argument("--work-dir", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path)
    args = parser.parse_args()
    work = args.work_dir.resolve()
    if not (work / ".article53-inputs-complete").is_file():
        raise FileNotFoundError("Run prepare_article53_transmission.py first")
    output = (args.output_dir or work / "summary").resolve()
    if output.exists():
        shutil.rmtree(output)
    output.mkdir(parents=True)

    design = mother_design(work, output)
    species = mother_species(work, output)
    strain_rows = published_mother_strains(work, output)
    fmt = fmt_outputs(work, output)
    evidence_ladder(output)
    summary = {
        "mother_infant_samples": species["samples"],
        "mother_infant_timepoint_sets": len(design),
        "mother_infant_families": len({row["Pair"] for row in design}),
        "mother_infant_named_species": species["named_species"],
        "primary_presence_threshold_pct_exclusive": 0.1,
        "primary_matched_median_shared_species": species["primary"]["MatchedMedianSharedSpecies"],
        "primary_unrelated_median_shared_species": species["primary"]["UnrelatedMedianSharedSpecies"],
        "primary_matched_median_jaccard": species["primary"]["MatchedMedianJaccard"],
        "primary_unrelated_median_jaccard": species["primary"]["UnrelatedMedianJaccard"],
        "primary_infants_with_negative_controls": species["primary"]["InfantsWithNegativeControls"],
        "primary_matched_ranked_first": species["primary"]["MatchedRankedFirstIncludingTies"],
        "quantitative_mother_infant_strain_events": len(strain_rows),
        "fmt_workbook_samples": fmt["workbook_samples"],
        "fmt_rcdi_samples": fmt["rcdi_samples"],
        "fmt_control_samples": fmt["control_samples"],
        "fmt_cases": fmt["cases"],
        "fmt_complete_paired_cases": fmt["complete_paired_cases"],
        "fmt_competing_strain_events": fmt["events"],
        "fmt_source_counts": fmt["source_counts"],
        "fmt_shared_strain_prepost_median": fmt["paired_summary"]["Strain"]["PrePostMedian"],
        "fmt_shared_strain_donorpost_median": fmt["paired_summary"]["Strain"]["DonorPostMedian"],
        "fmt_shared_species_prepost_median": fmt["paired_summary"]["Species"]["PrePostMedian"],
        "fmt_shared_species_donorpost_median": fmt["paired_summary"]["Species"]["DonorPostMedian"],
        "same_str_mvs_threshold": 0.999,
        "same_str_overlap_bp": 5000,
        "random_output_requested": False,
        "seed": 20260753,
    }
    dump_json(output / "summary.json", summary)
    (work / ".article53-summary-complete").write_text("complete\n", encoding="utf-8")
    print(
        "Summarized 24 mother-infant profiles, 27 FMT cases and "
        "408 competing-strain events"
    )


if __name__ == "__main__":
    main()
