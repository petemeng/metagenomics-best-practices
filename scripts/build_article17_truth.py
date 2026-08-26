#!/usr/bin/env python3
"""Freeze the Article 17 MOCK1 truth and NCBI accession snapshots.

This is a one-time provenance step. It transcribes the 71 MOCK1 rows from the
publisher XLSX without renormalizing the expected percentages, joins the
official NCBI Datasets genome report, and freezes deterministic JSONL copies.
Database-specific reference presence is computed later from each extracted
Kraken database rather than inferred from this table.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
from pathlib import Path
from typing import Any

import openpyxl


EXPECTED_XLSX_BYTES = 18_203
EXPECTED_XLSX_SHA256 = (
    "6bb36d2121ff74b0542620e1e15b96d83256c797576f6e9fb192929f1ec3c12f"
)
EXPECTED_ROWS = 71
EXPECTED_SUM = 100.01215654459679
RETRIEVAL_DATE = "2026-07-21"
DATASETS_VERSION = "18.33.1"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--supplement-xlsx", type=Path, required=True)
    parser.add_argument("--genome-jsonl", type=Path, required=True)
    parser.add_argument("--sequence-jsonl", type=Path, required=True)
    parser.add_argument("--truth-output", type=Path, required=True)
    parser.add_argument("--genome-output", type=Path, required=True)
    parser.add_argument("--sequence-output", type=Path, required=True)
    parser.add_argument("--provenance-output", type=Path, required=True)
    return parser.parse_args()


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line]


def write_jsonl(path: Path, rows: list[dict[str, Any]], key) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        for row in sorted(rows, key=key):
            handle.write(json.dumps(row, ensure_ascii=False, sort_keys=True, separators=(",", ":")))
            handle.write("\n")


def main() -> None:
    args = parse_args()
    if args.supplement_xlsx.stat().st_size != EXPECTED_XLSX_BYTES:
        raise SystemExit("publisher XLSX byte-count mismatch")
    if sha256(args.supplement_xlsx) != EXPECTED_XLSX_SHA256:
        raise SystemExit("publisher XLSX SHA-256 mismatch")

    workbook = openpyxl.load_workbook(args.supplement_xlsx, data_only=True, read_only=True)
    if workbook.sheetnames != ["summary"]:
        raise SystemExit(f"unexpected workbook sheets: {workbook.sheetnames}")
    sheet = workbook["summary"]
    truth_source: list[dict[str, Any]] = []
    for index, values in enumerate(
        sheet.iter_rows(min_row=2, max_row=72, min_col=1, max_col=4, values_only=True),
        start=1,
    ):
        organism, strain, accession, expected = values
        truth_source.append(
            {
                "TruthIndex": index,
                "ExpectedOrganismGTDBRS207": str(organism).strip(),
                "StrainName": str(strain).strip(),
                "AssemblyAccession": str(accession).strip(),
                "ExpectedGenomePercent": float(expected),
            }
        )
    if len(truth_source) != EXPECTED_ROWS:
        raise SystemExit(f"expected {EXPECTED_ROWS} truth rows, observed {len(truth_source)}")
    observed_sum = sum(row["ExpectedGenomePercent"] for row in truth_source)
    if abs(observed_sum - EXPECTED_SUM) > 1e-12:
        raise SystemExit(f"publisher abundance sum changed: {observed_sum!r}")

    genome_rows = read_jsonl(args.genome_jsonl)
    sequence_rows = read_jsonl(args.sequence_jsonl)
    genome_by_accession = {row["accession"]: row for row in genome_rows}
    sequence_by_accession: dict[str, list[dict[str, Any]]] = {}
    for row in sequence_rows:
        sequence_by_accession.setdefault(row["assembly_accession"], []).append(row)
    expected_accessions = {row["AssemblyAccession"] for row in truth_source}
    if set(genome_by_accession) != expected_accessions:
        missing = sorted(expected_accessions - set(genome_by_accession))
        extra = sorted(set(genome_by_accession) - expected_accessions)
        raise SystemExit(f"NCBI genome report identity mismatch; missing={missing}, extra={extra}")

    output_rows: list[dict[str, Any]] = []
    for source in truth_source:
        accession = source["AssemblyAccession"]
        genome = genome_by_accession[accession]
        organism = genome["organism"]
        assembly_info = genome.get("assembly_info", {})
        sequences = sequence_by_accession.get(accession, [])
        refseq_accessions = sorted(
            {row["refseq_accession"] for row in sequences if row.get("refseq_accession")}
        )
        genbank_accessions = sorted(
            {row["genbank_accession"] for row in sequences if row.get("genbank_accession")}
        )
        output_rows.append(
            {
                **source,
                "SourceDOI": "10.1038/s41597-022-01762-z",
                "SupplementaryTable": "S3",
                "PublisherXLSXBytes": EXPECTED_XLSX_BYTES,
                "PublisherXLSXSHA256": EXPECTED_XLSX_SHA256,
                "NCBIRetrievalDate": RETRIEVAL_DATE,
                "NCBIDatasetsVersion": DATASETS_VERSION,
                "CurrentAccession": genome.get("current_accession", ""),
                "PairedRefSeqAccession": genome.get("paired_accession", ""),
                "NCBITaxID": organism["tax_id"],
                "NCBIOrganismName": organism["organism_name"],
                "NCBIStrain": organism.get("infraspecific_names", {}).get("strain", ""),
                "AssemblyStatus": assembly_info.get("assembly_status", ""),
                "AssemblyLevel": assembly_info.get("assembly_level", ""),
                "SequenceReportRows": len(sequences),
                "RefSeqSequenceAccessions": ";".join(refseq_accessions),
                "GenBankSequenceAccessions": ";".join(genbank_accessions),
            }
        )

    fields = list(output_rows[0])
    args.truth_output.parent.mkdir(parents=True, exist_ok=True)
    with args.truth_output.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, delimiter="\t", lineterminator="\n")
        writer.writeheader()
        writer.writerows(output_rows)

    write_jsonl(args.genome_output, genome_rows, key=lambda row: row["accession"])
    write_jsonl(
        args.sequence_output,
        sequence_rows,
        key=lambda row: (
            row["assembly_accession"],
            row.get("refseq_accession", ""),
            row.get("genbank_accession", ""),
            row.get("sequence_name", ""),
        ),
    )

    provenance = [
        {
            "Asset": "publisher_supplement_s3",
            "SourcePath": args.supplement_xlsx.name,
            "SourceBytes": args.supplement_xlsx.stat().st_size,
            "SourceSHA256": sha256(args.supplement_xlsx),
            "FrozenPath": "not_stored_binary",
            "FrozenBytes": 0,
            "FrozenSHA256": "not_applicable",
            "Retrieved": RETRIEVAL_DATE,
            "Tool": f"openpyxl {openpyxl.__version__}",
        },
        {
            "Asset": "ncbi_genome_report",
            "SourcePath": args.genome_jsonl.name,
            "SourceBytes": args.genome_jsonl.stat().st_size,
            "SourceSHA256": sha256(args.genome_jsonl),
            "FrozenPath": args.genome_output.name,
            "FrozenBytes": args.genome_output.stat().st_size,
            "FrozenSHA256": sha256(args.genome_output),
            "Retrieved": RETRIEVAL_DATE,
            "Tool": f"NCBI Datasets {DATASETS_VERSION}",
        },
        {
            "Asset": "ncbi_sequence_report",
            "SourcePath": args.sequence_jsonl.name,
            "SourceBytes": args.sequence_jsonl.stat().st_size,
            "SourceSHA256": sha256(args.sequence_jsonl),
            "FrozenPath": args.sequence_output.name,
            "FrozenBytes": args.sequence_output.stat().st_size,
            "FrozenSHA256": sha256(args.sequence_output),
            "Retrieved": RETRIEVAL_DATE,
            "Tool": f"NCBI Datasets {DATASETS_VERSION}",
        },
        {
            "Asset": "mock1_truth_crosswalk",
            "SourcePath": "publisher_supplement_s3+ncbi_reports",
            "SourceBytes": 0,
            "SourceSHA256": "derived",
            "FrozenPath": args.truth_output.name,
            "FrozenBytes": args.truth_output.stat().st_size,
            "FrozenSHA256": sha256(args.truth_output),
            "Retrieved": RETRIEVAL_DATE,
            "Tool": "build_article17_truth.py",
        },
    ]
    with args.provenance_output.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=list(provenance[0]),
            delimiter="\t",
            lineterminator="\n",
        )
        writer.writeheader()
        writer.writerows(provenance)

    print(
        json.dumps(
            {
                "truth_rows": len(output_rows),
                "truth_abundance_percent": observed_sum,
                "genome_reports": len(genome_rows),
                "sequence_reports": len(sequence_rows),
                "assemblies_with_sequence_reports": len(sequence_by_accession),
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
