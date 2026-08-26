#!/usr/bin/env python3
"""Lock the 71-genome MOCK1 truth set and Meslier Supplementary Table S2."""

from __future__ import annotations

import argparse
import csv
import gzip
import hashlib
import json
import os
import subprocess
from collections import Counter
from pathlib import Path

import openpyxl


ALIASES = {
    "Chlorobaculum_tepidum_TLS": "Chlorobium_tepidum_TLS",
    "Trichormus_sp._PCC_7120": "Nostoc_sp._PCC_7120",
    "Sulfurisphaera_tokodaii_str._7": "Sulfolobus_tokodaii_str._7",
    "Phocaeicola_vulgatus_ATCC_8482": "Bacteroides_vulgatus_ATCC_8482",
    "Micromonospora_tropica_CNB-440": "Salinispora_tropica_CNB-440",
    "Micromonospora_arenicola_CNS-205": "Salinispora_arenicola_CNS-205",
    "Desulfovibrio_mercurii_ND_132": "Desulfovibrio_desulfuricans_ND132",
    "Bordetella_pertussis_RB50": "Bordetella_bronchiseptica_RB50",
    "Nitratidesulfovibrio_vulgaris_Hildenborough": "Desulfovibrio_vulgaris_Hildenborough",
}


def args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--benchmark-repo", type=Path, required=True)
    parser.add_argument("--supplement-s2", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    return parser.parse_args()


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(8 * 1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def fasta_inventory(path: Path) -> tuple[int, int, Counter[str]]:
    opener = gzip.open if path.suffix == ".gz" else Path.open
    records = bases = 0
    sequence = bytearray()
    sequence_hashes: Counter[str] = Counter()
    with opener(path, "rb") as handle:
        for raw in handle:
            if raw.startswith(b">"):
                if sequence:
                    sequence_hashes[hashlib.sha256(sequence).hexdigest()] += 1
                    sequence.clear()
                records += 1
            else:
                clean = raw.strip().upper()
                bases += len(clean)
                sequence.extend(clean)
    if sequence:
        sequence_hashes[hashlib.sha256(sequence).hexdigest()] += 1
    return records, bases, sequence_hashes


def number(value):
    if isinstance(value, str):
        value = value.replace(",", "").strip()
        parsed = float(value)
        return int(parsed) if parsed.is_integer() else parsed
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def write_tsv(path: Path, rows: list[dict], fields: list[str]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, delimiter="\t", lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    a = args()
    repo = a.benchmark_repo.resolve()
    output = a.output_dir.resolve()
    output.mkdir(parents=True, exist_ok=True)
    refs_dir = output / "references" / "MOCK1"
    refs_dir.mkdir(parents=True, exist_ok=True)

    s1 = repo / "script_r" / "Supplementary_Table_S1.xlsx"
    combined = repo / "reference" / "MOCK_001.fasta.gz"
    genome_dir = repo / "reference" / "all_genomes_listed"
    for path in (s1, combined, a.supplement_s2):
        if not path.is_file() or path.stat().st_size == 0:
            raise SystemExit(f"Missing official source: {path}")

    workbook = openpyxl.load_workbook(s1, read_only=True, data_only=True)
    rows = list(workbook.active.iter_rows(values_only=True))[1:]
    truth = []
    aggregate_records = aggregate_bases = 0
    aggregate_hashes: Counter[str] = Counter()
    for row in rows:
        abundance = row[14]
        if abundance in (None, "NP"):
            continue
        current_label = str(row[2]).strip()
        file_label = ALIASES.get(current_label, current_label)
        source = genome_dir / f"{file_label}.fna.gz"
        if not source.is_file():
            raise ValueError(f"No reference FASTA for {current_label}: {source}")
        records, bases, hashes = fasta_inventory(source)
        aggregate_records += records
        aggregate_bases += bases
        aggregate_hashes.update(hashes)
        link = refs_dir / f"{current_label}.fna.gz"
        if link.is_symlink():
            if link.resolve() != source.resolve():
                raise ValueError(f"Reference symlink points elsewhere: {link}")
        elif link.exists():
            raise ValueError(f"Reference path exists but is not a symlink: {link}")
        else:
            os.symlink(source.resolve(), link)
        truth.append(
            {
                "CurrentGenomeLabel": current_label,
                "RepositoryFileLabel": file_label,
                "GenBankAssembly": str(row[3]).strip(),
                "ExpectedAbundancePct": f"{float(abundance):.12g}",
                "ReferenceRecords": records,
                "ReferenceBases": bases,
                "CompressedBytes": source.stat().st_size,
                "CompressedSHA256": sha256(source),
                "TaxonomyRenameAlias": "yes" if current_label != file_label else "no",
            }
        )
    truth.sort(key=lambda item: item["CurrentGenomeLabel"])
    if len(truth) != 71:
        raise ValueError(f"Expected 71 MOCK1 genomes, found {len(truth)}")
    expected_links = {
        f"{row['CurrentGenomeLabel']}.fna.gz" for row in truth
    }
    observed_paths = list(refs_dir.glob("*.fna.gz"))
    observed_links = {path.name for path in observed_paths}
    if observed_links != expected_links or not all(
        path.is_symlink() for path in observed_paths
    ):
        raise ValueError(
            "MOCK1 reference symlink inventory differs from the 71-member truth set"
        )
    combined_records, combined_bases, combined_hashes = fasta_inventory(combined)
    if (aggregate_records, aggregate_bases, aggregate_hashes) != (
        combined_records,
        combined_bases,
        combined_hashes,
    ):
        raise ValueError("The 71 selected reference files do not equal MOCK_001.fasta.gz")
    abundance_sum = sum(float(row["ExpectedAbundancePct"]) for row in truth)
    if abs(abundance_sum - 100) > 0.1:
        raise ValueError(f"MOCK1 expected abundances sum to {abundance_sum}, not approximately 100%")
    write_tsv(
        output / "truth-manifest.tsv",
        truth,
        [
            "CurrentGenomeLabel",
            "RepositoryFileLabel",
            "GenBankAssembly",
            "ExpectedAbundancePct",
            "ReferenceRecords",
            "ReferenceBases",
            "CompressedBytes",
            "CompressedSHA256",
            "TaxonomyRenameAlias",
        ],
    )

    s2book = openpyxl.load_workbook(a.supplement_s2, read_only=True, data_only=True)
    sheet = s2book.active
    values = list(sheet.iter_rows(values_only=True))
    fields = {
        "ReadCount": "Nb Reads (M)",
        "Contigs": "Nb Contigs",
        "LargestContigBp": "Largest Contig (bp)",
        "N50Bp": "N50 (bp)",
        "GenomeFractionPct": "Genome Fraction(%)",
        "MismatchesPer100Kbp": "Mismatches per 100kbps",
        "IndelsPer100Kbp": "Indels Per 100kbps",
        "FullyUnalignedContigs": "Fully Unaligned Contigs",
        "FullyUnalignedLengthBp": "Fully Unaligned Length (bp)",
        "FullGenomes99Pct": "NB full genome*",
    }

    def block(start: int) -> dict[str, tuple]:
        return {str(row[0]).strip(): row for row in values[start : start + 12] if row[0] is not None}

    ont, hifi = block(1), block(18)
    branches = [
        ("Published Illumina-only", ont, 5, "short-only", "SPAdes 3.14.1"),
        ("Published Illumina+ONT", ont, 6, "short-read-first hybrid", "SPAdes 3.14.1"),
        ("Published ONT-only", ont, 11, "long-only", "metaFlye 2.8.1"),
        ("Published Illumina+HiFi", hifi, 6, "short-read-first hybrid", "SPAdes 3.14.1"),
        ("Published HiFi-only", hifi, 11, "long-only", "metaFlye 2.8.1"),
    ]
    anchor = []
    for label, source, column, mode, tool in branches:
        item = {"Branch": label, "Mode": mode, "Tool": tool}
        for output_name, source_name in fields.items():
            value = source[source_name][column]
            item[output_name] = str(value) if output_name == "ReadCount" else number(value)
        anchor.append(item)
    write_tsv(output / "published-anchor.tsv", anchor, ["Branch", "Mode", "Tool", *fields])

    commit = subprocess.check_output(
        ["git", "-C", str(repo), "rev-parse", "HEAD"], text=True
    ).strip()
    audit = {
        "benchmark_repository": "https://forgemia.inra.fr/metagenopolis/benchmark_mock.git",
        "benchmark_commit": commit,
        "supplement_s1_sha256": sha256(s1),
        "supplement_s2_sha256": sha256(a.supplement_s2),
        "combined_reference_sha256": sha256(combined),
        "mock1_genomes": len(truth),
        "mock1_reference_records": combined_records,
        "mock1_reference_bases": combined_bases,
        "mock1_expected_abundance_sum_pct": abundance_sum,
        "taxonomy_rename_aliases": len(ALIASES),
        "sequence_multiset_equal": True,
    }
    (output / "truth-audit.json").write_text(json.dumps(audit, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(audit))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
