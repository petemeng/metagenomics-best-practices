#!/usr/bin/env python3
"""Download and verify the version-locked reporting-standard sources for Article 76."""

from __future__ import annotations

import argparse
import hashlib
import json
import tempfile
import urllib.request
import xml.etree.ElementTree as ET
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image


ARTICLE = 76
SNAPSHOT_DATE = "2026-08-23"
SOURCES = {
    "STREAMS_Guidelines_Zenodo.xlsx": {
        "url": "https://zenodo.org/api/records/15014818/files/STREAMS_Guidelines_Zenodo.xlsx/content",
        "bytes": 398352,
        "sha256": "c3905dbdd28f256968a06157682277a50712ad045b9624a0f11cce47414ff788",
        "md5": "911e3eb733e1069fc4d7e40d7b2edefe",
        "identity": "STREAMS Guidelines v1.0; Zenodo 10.5281/zenodo.15014818",
    },
    "STORMS_Excel_1.03.xlsx": {
        "url": "https://zenodo.org/api/records/5714305/files/STORMS_Excel_1.03.xlsx/content",
        "bytes": 76465,
        "sha256": "4d763f0d62ee27aa43217a0ebe47e75b9beb70c8534f867c5c2fc3b6f1714b60",
        "md5": "4962fd585c25bd797c1319e648c46d0a",
        "identity": "STORMS checklist v1.03; Zenodo 10.5281/zenodo.5714305",
    },
    "PMC6436528-fulltext.xml": {
        "url": "https://www.ebi.ac.uk/europepmc/webservices/rest/PMC6436528/fullTextXML",
        "bytes": 175407,
        "sha256": "7fdec05ef70b15e852d21225fbcae8b0ecc24e4accdaf5075d1e976ab9295ba4",
        "identity": "MIMAG; DOI 10.1038/nbt.3893",
    },
    "PMC6871006-fulltext.xml": {
        "url": "https://www.ebi.ac.uk/europepmc/webservices/rest/PMC6871006/fullTextXML",
        "bytes": 183966,
        "sha256": "289c221ece18f17a0a8930a354cfc3e01b712751abfdd52ff7b054c6f31c36fa",
        "identity": "MIUViG; DOI 10.1038/nbt.4306",
    },
    "streams-figure1-original.png": {
        "url": "https://media.springernature.com/full/springer-static/image/art%3A10.1038%2Fs41564-025-02186-2/MediaObjects/41564_2025_2186_Fig1_HTML.png",
        "bytes": 299584,
        "sha256": "993fb497cc09e2e3446b22233ca70ab9bacff9042babce39955251288a34279d",
        "identity": "STREAMS Figure 1; 2174 x 1098 px",
    },
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--verify-only", action="store_true")
    return parser.parse_args()


def digest(path: Path, algorithm: str = "sha256") -> str:
    checksum = hashlib.new(algorithm)
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(8 * 1024 * 1024), b""):
            checksum.update(block)
    return checksum.hexdigest()


def valid(path: Path, record: dict[str, object]) -> bool:
    if not path.is_file() or path.stat().st_size != record["bytes"]:
        return False
    if digest(path) != record["sha256"]:
        return False
    return "md5" not in record or digest(path, "md5") == record["md5"]


def download(url: str, destination: Path) -> None:
    request = urllib.request.Request(
        url, headers={"User-Agent": "metagenomics-best-practices/article-76"}
    )
    with tempfile.NamedTemporaryFile(dir=destination.parent, delete=False) as handle:
        temporary = Path(handle.name)
        try:
            with urllib.request.urlopen(request, timeout=180) as response:
                while block := response.read(8 * 1024 * 1024):
                    handle.write(block)
        except Exception:
            temporary.unlink(missing_ok=True)
            raise
    temporary.replace(destination)


def workbook_semantics(output: Path) -> None:
    streams = load_workbook(
        output / "STREAMS_Guidelines_Zenodo.xlsx", read_only=True, data_only=True
    )
    if "STREAMS_final" not in streams.sheetnames:
        raise ValueError("STREAMS v1.0 workbook lacks STREAMS_final")
    numeric = [
        row[0]
        for row in streams["STREAMS_final"].iter_rows(min_row=2, values_only=True)
        if isinstance(row[0], (int, float)) and not isinstance(row[0], bool)
    ]
    if len(numeric) != 67:
        raise ValueError(f"Expected 67 STREAMS recommendations, observed {len(numeric)}")

    storms = load_workbook(
        output / "STORMS_Excel_1.03.xlsx", read_only=True, data_only=True
    )
    if storms["Checklist"]["B1"].value != 1.03:
        raise ValueError("Unexpected STORMS checklist version")
    top_level = [
        row[0]
        for row in storms["Checklist"].iter_rows(min_row=3, values_only=True)
        if isinstance(row[0], (int, float))
        and not isinstance(row[0], bool)
        and float(row[0]).is_integer()
    ]
    if top_level != list(range(1, 18)):
        raise ValueError(f"Unexpected STORMS top-level items: {top_level}")


def xml_semantics(output: Path) -> None:
    expected = {
        "PMC6436528-fulltext.xml": "10.1038/nbt.3893",
        "PMC6871006-fulltext.xml": "10.1038/nbt.4306",
    }
    for filename, doi in expected.items():
        root = ET.parse(output / filename).getroot()
        observed = root.findtext('.//article-id[@pub-id-type="doi"]')
        if observed != doi:
            raise ValueError(f"Unexpected DOI in {filename}: {observed!r}")
    with Image.open(output / "streams-figure1-original.png") as image:
        if image.size != (2174, 1098):
            raise ValueError(f"Unexpected STREAMS Figure 1 dimensions: {image.size}")


def main() -> None:
    args = parse_args()
    output = args.output_dir.resolve()
    output.mkdir(parents=True, exist_ok=True)
    for filename, record in SOURCES.items():
        path = output / filename
        if not valid(path, record):
            if args.verify_only:
                raise RuntimeError(f"Missing or checksum-mismatched source: {path}")
            download(str(record["url"]), path)
        if not valid(path, record):
            raise RuntimeError(f"Downloaded source failed byte lock: {path}")
        print(f"verified\t{filename}\t{record['sha256']}")

    workbook_semantics(output)
    xml_semantics(output)
    manifest = {
        "article": ARTICLE,
        "snapshot_date": SNAPSHOT_DATE,
        "sources": SOURCES,
        "workbook_policy": (
            "Parse STREAMS_final only. Other workbook sheets include development "
            "material and are not part of the published v1.0 checklist used here."
        ),
    }
    (output / "download-manifest.json").write_text(
        json.dumps(manifest, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )
    print(f"manifest\t{output / 'download-manifest.json'}")


if __name__ == "__main__":
    main()
