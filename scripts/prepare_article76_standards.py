#!/usr/bin/env python3
"""Build Article 76's reporting-standard crosswalk and real-data audits."""

from __future__ import annotations

import argparse
import hashlib
import json
import platform
import shutil
import xml.etree.ElementTree as ET
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from PIL import Image


ARTICLE = 76
ANALYSIS_SEED = 76_001
PLOT_SEED = 20_260_776
SNAPSHOT_DATE = "2026-08-23"
MANUSCRIPT_SECTIONS = (
    "Abstract",
    "Introduction",
    "Methods",
    "Results",
    "Discussion",
    "Other information",
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--project-root", type=Path, required=True)
    parser.add_argument("--evidence-dir", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    return parser.parse_args()


def sha256(path: Path) -> str:
    checksum = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(8 * 1024 * 1024), b""):
            checksum.update(block)
    return checksum.hexdigest()


def write_tsv(frame: pd.DataFrame, path: Path) -> None:
    frame.to_csv(path, sep="\t", index=False, lineterminator="\n", float_format="%.10g")


def normalized(node: ET.Element | None) -> str:
    return "" if node is None else " ".join("".join(node.itertext()).split())


def item_number(cell: object) -> str:
    value = cell.value
    if isinstance(value, int) or (isinstance(value, float) and value.is_integer()):
        return str(int(value))
    if not isinstance(value, float):
        return str(value)
    fmt = str(cell.number_format)
    decimals = len(fmt.split(".", 1)[1]) if "." in fmt and set(fmt) <= {"0", "."} else 1
    return f"{value:.{decimals}f}"


def parse_checklist(path: Path, sheet: str, standard: str, first_row: int) -> pd.DataFrame:
    workbook = load_workbook(path, read_only=False, data_only=True)
    worksheet = workbook[sheet]
    section = ""
    rows: list[dict[str, object]] = []
    for row_number in range(first_row, worksheet.max_row + 1):
        first = worksheet.cell(row_number, 1)
        value = first.value
        if isinstance(value, str) and value.strip():
            section = " ".join(value.split())
            continue
        if not isinstance(value, (int, float)) or isinstance(value, bool):
            continue
        rows.append(
            {
                "Standard": standard,
                "WorkbookSheet": sheet,
                "SourceRow": row_number,
                "ManuscriptSection": section,
                "ItemNumber": item_number(first),
                "TopLevelItem": int(float(value)),
                "Item": " ".join(str(worksheet.cell(row_number, 2).value or "").split()),
                "Recommendation": " ".join(str(worksheet.cell(row_number, 3).value or "").split()),
                "ItemSource": " ".join(str(worksheet.cell(row_number, 4).value or "").split()),
                "AdditionalGuidance": " ".join(str(worksheet.cell(row_number, 5).value or "").split()),
            }
        )
    frame = pd.DataFrame(rows)
    if not set(frame["ManuscriptSection"]).issubset(MANUSCRIPT_SECTIONS):
        raise ValueError(f"Unexpected {standard} section labels")
    return frame


def table_rows(xml_path: Path, number: int) -> list[list[str]]:
    root = ET.parse(xml_path).getroot()
    tables = root.findall(".//table-wrap")
    if number < 1 or number > len(tables):
        raise ValueError(f"Table {number} not found in {xml_path}")
    rows: list[list[str]] = []
    for row in tables[number - 1].findall(".//tr"):
        cells = [
            normalized(cell)
            for cell in list(row)
            if cell.tag.rsplit("}", 1)[-1] in {"td", "th"}
        ]
        if cells:
            rows.append(cells)
    return rows


def parse_mimag(xml_path: Path) -> pd.DataFrame:
    rows = table_rows(xml_path, 1)[1:]
    current = ""
    output = []
    for row in rows:
        if len(row) == 1:
            current = row[0].replace(" (SAG/MAG)", "")
            continue
        output.append({"QualityLevel": current, "Criterion": row[0], "Requirement": row[1]})
    frame = pd.DataFrame(output)
    if set(frame["QualityLevel"]) != {
        "Finished",
        "High-quality draft",
        "Medium-quality draft",
        "Low-quality draft",
    }:
        raise ValueError("Unexpected MIMAG quality levels")
    return frame


def parse_miuvig(xml_path: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    mandatory_rows = table_rows(xml_path, 1)[1:]
    mandatory = pd.DataFrame(mandatory_rows, columns=["MandatoryMetadata", "Description"])
    if len(mandatory) != 8:
        raise ValueError(f"Expected eight MIUViG mandatory fields, observed {len(mandatory)}")

    rows = table_rows(xml_path, 2)
    header = rows[0]
    output = []
    for criterion_row in rows[1:]:
        for category, requirement in zip(header[1:], criterion_row[1:], strict=True):
            output.append(
                {
                    "Category": category,
                    "Criterion": criterion_row[0],
                    "Requirement": requirement,
                }
            )
    categories = pd.DataFrame(output)
    return mandatory, categories


def selection_matrix() -> pd.DataFrame:
    scenarios = {
        "Human · read profiles": ("Core", "Not selected", "Not selected", "Not selected"),
        "Human · MAGs": ("Core", "Not selected", "Add", "Not selected"),
        "Human · UViGs": ("Core", "Not selected", "Not selected", "Add"),
        "Environmental · read profiles": ("Not selected", "Core", "Not selected", "Not selected"),
        "Environmental · MAGs": ("Not selected", "Core", "Add", "Not selected"),
        "Non-human host · MAGs + UViGs": ("Not selected", "Core", "Add", "Add"),
        "Synthetic community · UViGs": ("Not selected", "Core", "Not selected", "Add"),
    }
    rows = []
    for scenario, states in scenarios.items():
        for standard, state in zip(("STORMS", "STREAMS", "MIMAG", "MIUViG"), states, strict=True):
            rows.append(
                {
                    "Scenario": scenario,
                    "Standard": standard,
                    "Role": state,
                    "RoleCode": {"Not selected": 0, "Add": 1, "Core": 2}[state],
                }
            )
    return pd.DataFrame(rows)


def layer_map() -> pd.DataFrame:
    rows = [
        ("Study and manuscript", "STORMS", "Core", "Human design, participants, ethics, results and disclosure"),
        ("Study and manuscript", "STREAMS", "Core", "Environmental, non-human host or synthetic study context"),
        ("Study and manuscript", "MIMAG", "Entity add-on", "Does not replace a study-level checklist"),
        ("Study and manuscript", "MIUViG", "Entity add-on", "Does not replace a study-level checklist"),
        ("Sample and laboratory", "STORMS", "Core", "Collection, storage, extraction, controls and batches"),
        ("Sample and laboratory", "STREAMS", "Core", "Spatial-temporal context, preservation, controls and quantities"),
        ("Sample and laboratory", "MIMAG", "Supporting", "Source and sequencing context enter the genome record"),
        ("Sample and laboratory", "MIUViG", "Supporting", "Dataset source and detection type enter the UViG record"),
        ("Analysis and results", "STORMS", "Core", "QC, software, databases, statistics and findings"),
        ("Analysis and results", "STREAMS", "Core", "Bioinformatics, bias, missingness and results"),
        ("Analysis and results", "MIMAG", "Core for MAGs", "Assembly, completeness, contamination, rRNA and tRNA"),
        ("Analysis and results", "MIUViG", "Core for UViGs", "Identification, structure, completeness and annotation"),
        ("Entity and archive", "STORMS", "Supporting", "Raw, processed, participant and code access"),
        ("Entity and archive", "STREAMS", "Supporting", "Metadata, raw/processed data, code and availability"),
        ("Entity and archive", "MIMAG", "Core for MAGs", "One quality record per SAG/MAG"),
        ("Entity and archive", "MIUViG", "Core for UViGs", "One mandatory-metadata record per UViG"),
    ]
    return pd.DataFrame(rows, columns=["Layer", "Standard", "Role", "Record"])


def crosswalk() -> pd.DataFrame:
    levels = {
        "Design and context": ("Core", "Core", "Supporting", "Supporting"),
        "Sampling and metadata": ("Core", "Core", "Supporting", "Core"),
        "Laboratory controls": ("Core", "Core", "Not primary", "Not primary"),
        "Bioinformatics and statistics": ("Core", "Core", "Supporting", "Supporting"),
        "Software and database versions": ("Core", "Core", "Core", "Core"),
        "Data and code access": ("Core", "Core", "Supporting", "Supporting"),
        "Bacterial/archaeal genome quality": ("Supporting", "Supporting", "Core", "Not primary"),
        "Virus genome quality": ("Supporting", "Supporting", "Not primary", "Core"),
        "Results, limits and generalizability": ("Core", "Core", "Not primary", "Not primary"),
        "Ethics, consent and governance": ("Core", "Context-specific", "Not primary", "Not primary"),
    }
    rows = []
    for domain, values in levels.items():
        for standard, coverage in zip(("STORMS", "STREAMS", "MIMAG", "MIUViG"), values, strict=True):
            rows.append(
                {
                    "ReportingDomain": domain,
                    "Standard": standard,
                    "Coverage": coverage,
                    "CoverageCode": {
                        "Not primary": 0,
                        "Supporting": 1,
                        "Context-specific": 1,
                        "Core": 2,
                    }[coverage],
                }
            )
    return pd.DataFrame(rows)


def mimag_audit(root: Path) -> pd.DataFrame:
    source = root / "data/small/44-mag-qc-mimag-graph-frozen/mag-quality-summary.tsv"
    frame = pd.read_csv(source, sep="\t")
    booleans = ("GUNCPass", "Complete5S", "Complete16S", "Complete23S", "CompleteRRNASet")
    for column in booleans:
        frame[column] = frame[column].astype(str).str.lower().map({"true": True, "false": False})
    core_hq = (
        frame["CheckM2Completeness"].gt(90)
        & frame["CheckM2Contamination"].lt(5)
        & frame["CompleteRRNASet"]
        & frame["TRNAIsotypes"].ge(18)
    )
    core_mq = frame["CheckM2Completeness"].ge(50) & frame["CheckM2Contamination"].lt(10)
    frame["MIMAGCoreTier"] = "Low/failed"
    frame.loc[core_mq, "MIMAGCoreTier"] = "Medium quality"
    frame.loc[core_hq, "MIMAGCoreTier"] = "High quality"
    frame["Article44ExtendedTier"] = frame["MIMAGQuality"]
    frame["CoreVsExtendedAgreement"] = frame["MIMAGCoreTier"].eq(frame["Article44ExtendedTier"])
    frame["HQMissingGate"] = "No missing HQ gate"
    candidates = frame["CheckM2Completeness"].gt(90) & frame["CheckM2Contamination"].lt(5)
    frame.loc[candidates & ~frame["CompleteRRNASet"], "HQMissingGate"] = "Complete 5S/16S/23S set"
    frame.loc[candidates & frame["CompleteRRNASet"] & frame["TRNAIsotypes"].lt(18), "HQMissingGate"] = "At least 18 tRNA isotypes"
    columns = [
        "MAG",
        "CheckM2Completeness",
        "CheckM2Contamination",
        "GUNCPass",
        "Complete5S",
        "Complete16S",
        "Complete23S",
        "CompleteRRNASet",
        "TRNAIsotypes",
        "MIMAGCoreTier",
        "Article44ExtendedTier",
        "CoreVsExtendedAgreement",
        "HQMissingGate",
    ]
    return frame[columns]


def miuvig_audit(root: Path) -> pd.DataFrame:
    source = "Article 54 frozen fixture"
    rows = [
        ("Source of UViGs", "Complete", "input-lineage.tsv identifies 46 CheckV regression-fixture sequences"),
        ("Assembly software", "Missing", "Upstream assembly program and parameters are not carried by the public fixture"),
        ("Virus identification software", "Complete", "geNomad 1.12.0 and VirSorter2 2.2.4 versions, commits, parameters and logs"),
        ("Predicted genome type", "Missing", "DNA/RNA and strandedness are not recorded per UViG"),
        ("Predicted genome structure", "Partial", "terminal-repeat evidence exists, but structure is not resolved for every UViG"),
        ("Detection type", "Complete", "computational detection is recorded in the two-caller evidence matrix"),
        ("Assembly quality", "Complete", "CheckV 1.1.1 quality category is recorded for all 46 sequences"),
        ("Number of contigs", "Partial", "each record is one sequence, but segmented-genome membership is unresolved"),
    ]
    frame = pd.DataFrame(rows, columns=["MandatoryMetadata", "Status", "EvidenceOrGap"])
    frame.insert(0, "AuditObject", source)
    frame["StatusCode"] = frame["Status"].map({"Missing": 0, "Partial": 1, "Complete": 2})
    return frame


def responsibilities() -> pd.DataFrame:
    rows = [
        (1, "Protocol freeze", "Study lead", "Design, hypotheses, inclusion and sampling frame", "STORMS/STREAMS", "Before collection"),
        (2, "Field sampling", "Field team", "Site, time, environment/host and preservation ledger", "STREAMS", "At collection"),
        (3, "Participant governance", "Clinical team", "Consent, ethics, protected-data access path", "STORMS", "Before recruitment"),
        (4, "Wet lab", "Laboratory lead", "Kits/lots, controls, extraction, library and batch map", "STORMS/STREAMS", "At library freeze"),
        (5, "Sequence archive", "Data steward", "BioProject/BioSample/SRA accessions and ID crosswalk", "STORMS/STREAMS", "Before submission"),
        (6, "Bioinformatics", "Pipeline lead", "Commands, versions, databases, parameters and checksums", "All four", "At run freeze"),
        (7, "MAG ledger", "Genome analyst", "Per-MAG quality and feature record", "MIMAG", "Before MAG claims"),
        (8, "UViG ledger", "Virus analyst", "Per-UViG source, detection, structure and quality", "MIUViG", "Before UViG claims"),
        (9, "Statistics", "Statistical lead", "Units, models, missingness, multiplicity and sensitivity", "STORMS/STREAMS", "Before figure freeze"),
        (10, "Availability", "Data steward", "Processed data, code DOI, licenses and restrictions", "STORMS/STREAMS", "Before submission"),
        (11, "N/A adjudication", "Corresponding author", "Reason and approver for every not-applicable item", "All four", "At checklist freeze"),
        (12, "Final sign-off", "Corresponding author", "Versioned checklist plus manuscript page/line links", "All four", "At submission"),
    ]
    return pd.DataFrame(rows, columns=["Order", "Milestone", "Owner", "RequiredArtifact", "Standards", "Due"])


def na_ledger() -> pd.DataFrame:
    rows = [
        ("Human participant consent", "STORMS", "Not applicable", "Worked example uses public microbial benchmark sequences and has no participants", "Study lead"),
        ("Host genome access", "STREAMS 8.1", "Not applicable", "No host genome was generated or used", "Data steward"),
        ("Metatranscriptomics", "STORMS/STREAMS", "Not applicable", "DNA shotgun data only; no RNA library", "Laboratory lead"),
        ("Metabolomics/proteomics", "STORMS/STREAMS", "Not applicable", "No linked metabolite or protein assay", "Study lead"),
        ("Finished MAG claim", "MIMAG", "Not applicable", "No single-contig Q50-equivalent assembly with manual finishing", "Genome analyst"),
        ("Finished UViG claim", "MIUViG", "Not applicable", "Terminal repeats alone do not establish comprehensive manual finishing", "Virus analyst"),
        ("Causal inference", "STORMS/STREAMS", "Not applicable", "The worked audit makes reporting and quality claims, not causal claims", "Statistical lead"),
        ("AI usage", "STREAMS 18", "Report", "Record tool, version and exact task if AI contributed to the submitted work", "Corresponding author"),
    ]
    frame = pd.DataFrame(rows, columns=["Field", "StandardItem", "Disposition", "Reason", "Approver"])
    frame["ReasonRecorded"] = frame["Reason"].str.len().ge(25)
    return frame


def readiness(miuvig: pd.DataFrame, mimag: pd.DataFrame) -> pd.DataFrame:
    complete_mimag = bool(mimag["CoreVsExtendedAgreement"].all())
    rows = [
        ("Study checklist", "STREAMS", "Complete", "Published v1.0 sheet selected; all 67 recommendations indexed"),
        ("Study checklist", "STORMS", "Not selected", "Worked scenario is environmental/non-human, not human"),
        ("Sample metadata", "STREAMS", "Missing", "Tutorial fixture lacks a manuscript-level sampling table"),
        ("Positive/negative controls", "STREAMS", "Partial", "Benchmark truth exists; experimental controls are not applicable to the fixture"),
        ("Software/database versions", "All four", "Complete", "Version and database ledgers exist in Articles 44 and 54"),
        ("Raw/processed accessions", "STREAMS", "Partial", "Public source identities exist; manuscript accessions remain to be assigned"),
        ("Completeness/contamination", "MIMAG", "Complete", "23 MAGs have per-genome CheckM2 records"),
        ("rRNA/tRNA inventory", "MIMAG", "Complete", "23 MAGs have complete marker ledgers"),
        ("Chimera extension", "MIMAG + GUNC", "Complete", "GUNC pass and clade-separation fields are retained"),
        ("MIMAG tier reproducibility", "MIMAG", "Complete" if complete_mimag else "Partial", "Core and extended tiers compared per MAG"),
    ]
    for row in miuvig.itertuples(index=False):
        rows.append((row.MandatoryMetadata, "MIUViG", row.Status, row.EvidenceOrGap))
    rows.extend(
        [
            ("Not-applicable reasons", "All four", "Complete", "Every N/A entry has a reason and named approver"),
            ("Checklist page/line links", "STORMS/STREAMS", "Missing", "Add after manuscript pagination is stable"),
            ("AI contribution statement", "STREAMS 18", "Pending", "Apply current journal policy at submission"),
        ]
    )
    frame = pd.DataFrame(rows, columns=["ReadinessField", "Standard", "Status", "EvidenceOrNextAction"])
    frame["StatusCode"] = frame["Status"].map(
        {"Missing": 0, "Pending": 0, "Partial": 1, "Not selected": 1, "Complete": 2}
    )
    return frame


def source_manifest(evidence: Path, output: Path) -> dict[str, object]:
    manifest = json.loads((evidence / "download-manifest.json").read_text(encoding="utf-8"))
    if manifest.get("article") != ARTICLE:
        raise ValueError("Unexpected download manifest")
    identities = {
        "STREAMS_Guidelines_Zenodo.xlsx": ("STREAMS v1.0 checklist", "10.5281/zenodo.15014818", "CC BY 4.0 data record"),
        "STORMS_Excel_1.03.xlsx": ("STORMS v1.03 checklist", "10.5281/zenodo.5714305", "Zenodo record terms"),
        "PMC6436528-fulltext.xml": ("MIMAG full-text XML", "10.1038/nbt.3893", "Europe PMC full-text record"),
        "PMC6871006-fulltext.xml": ("MIUViG full-text XML", "10.1038/nbt.4306", "CC BY 4.0 article"),
        "streams-figure1-original.png": ("STREAMS Figure 1", "10.1038/s41564-025-02186-2", "Publisher/rightsholder terms; commentary only"),
    }
    records = []
    for filename, record in manifest["sources"].items():
        path = evidence / filename
        if not path.is_file() or path.stat().st_size != record["bytes"] or sha256(path) != record["sha256"]:
            raise ValueError(f"Source checksum mismatch: {filename}")
        title, doi, rights = identities[filename]
        records.append(
            {
                "File": filename,
                "Title": title,
                "DOI": doi,
                "URL": record["url"],
                "Bytes": record["bytes"],
                "SHA256": record["sha256"],
                "RightsBoundary": rights,
            }
        )
    anchor = output / "streams-figure1-original.png"
    shutil.copy2(evidence / "streams-figure1-original.png", anchor)
    with Image.open(anchor) as image:
        dimensions = {"width": image.width, "height": image.height}
    result = {
        "article": ARTICLE,
        "snapshot_date": SNAPSHOT_DATE,
        "records": records,
        "streams_sheet": "STREAMS_final",
        "excluded_streams_sheets": ["STREAMS", "STREAMS v2.0", "STREAMS_simplified1", "Group 4"],
        "anchor": {
            "file": anchor.name,
            "sha256": sha256(anchor),
            **dimensions,
            "rights_boundary": (
                "The original STREAMS Figure 1 is reproduced only for attributed "
                "scholarly commentary and is excluded from the repository CC BY/MIT grant."
            ),
        },
    }
    (output / "source-manifest.json").write_text(
        json.dumps(result, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )
    return result


def source_artifacts(root: Path, output: Path) -> None:
    paths = {
        "Article44_MAG_quality": root / "data/small/44-mag-qc-mimag-graph-frozen/mag-quality-summary.tsv",
        "Article44_MIMAG_counts": root / "data/small/44-mag-qc-mimag-graph-frozen/mimag-tier-counts.tsv",
        "Article54_input_lineage": root / "data/small/54-virus-discovery-quality-frozen/input-lineage.tsv",
        "Article54_tool_versions": root / "data/small/54-virus-discovery-quality-frozen/tool-versions.tsv",
        "Article54_CheckV_counts": root / "data/small/54-virus-discovery-quality-frozen/checkv-quality-counts.tsv",
        "Article54_terminal_repeats": root / "data/small/54-virus-discovery-quality-frozen/terminal-repeat-audit.tsv",
        "Article54_assertions": root / "data/small/54-virus-discovery-quality-frozen/miuvig-source-assertions.tsv",
        "Article54_summary": root / "data/small/54-virus-discovery-quality-frozen/summary.json",
    }
    records = []
    for key, path in paths.items():
        if not path.is_file():
            raise FileNotFoundError(path)
        records.append(
            {
                "ArtifactID": key,
                "RelativePath": path.relative_to(root).as_posix(),
                "Bytes": path.stat().st_size,
                "SHA256": sha256(path),
            }
        )
    write_tsv(pd.DataFrame(records), output / "source-artifact-manifest.tsv")


def main() -> None:
    args = parse_args()
    root = args.project_root.resolve()
    evidence = args.evidence_dir.resolve()
    output = args.output_dir.resolve()
    output.mkdir(parents=True, exist_ok=True)

    sources = source_manifest(evidence, output)
    storms = parse_checklist(evidence / "STORMS_Excel_1.03.xlsx", "Checklist", "STORMS", 3)
    streams = parse_checklist(evidence / "STREAMS_Guidelines_Zenodo.xlsx", "STREAMS_final", "STREAMS", 2)
    if len(storms) != 69 or storms["TopLevelItem"].nunique() != 17:
        raise ValueError("Unexpected STORMS checklist structure")
    if len(streams) != 67 or streams["TopLevelItem"].nunique() != 18:
        raise ValueError("Unexpected STREAMS checklist structure")
    checklists = pd.concat([storms, streams], ignore_index=True)
    write_tsv(checklists, output / "checklist-items.tsv")
    section_counts = (
        checklists.groupby(["Standard", "ManuscriptSection"], sort=False)
        .size()
        .rename("ExpandedRecommendations")
        .reset_index()
    )
    section_counts["ManuscriptSection"] = pd.Categorical(
        section_counts["ManuscriptSection"], categories=MANUSCRIPT_SECTIONS, ordered=True
    )
    section_counts = section_counts.sort_values(["Standard", "ManuscriptSection"])
    section_counts["ManuscriptSection"] = section_counts["ManuscriptSection"].astype(str)
    write_tsv(section_counts, output / "checklist-section-counts.tsv")

    mimag = parse_mimag(evidence / "PMC6436528-fulltext.xml")
    miuvig_fields, miuvig_categories = parse_miuvig(evidence / "PMC6871006-fulltext.xml")
    write_tsv(mimag, output / "mimag-quality-criteria.tsv")
    write_tsv(miuvig_fields, output / "miuvig-mandatory-metadata.tsv")
    write_tsv(miuvig_categories, output / "miuvig-quality-categories.tsv")

    selection = selection_matrix()
    layers = layer_map()
    domains = crosswalk()
    mag = mimag_audit(root)
    virus = miuvig_audit(root)
    owners = responsibilities()
    nas = na_ledger()
    ready = readiness(virus, mag)
    write_tsv(selection, output / "standard-selection-matrix.tsv")
    write_tsv(layers, output / "reporting-layer-map.tsv")
    write_tsv(domains, output / "standards-crosswalk.tsv")
    write_tsv(mag, output / "article44-mimag-compliance.tsv")
    write_tsv(virus, output / "article54-miuvig-compliance.tsv")
    write_tsv(owners, output / "field-responsibility-matrix.tsv")
    write_tsv(nas, output / "not-applicable-ledger.tsv")
    write_tsv(ready, output / "submission-readiness.tsv")
    source_artifacts(root, output)

    status_counts = ready["Status"].value_counts().to_dict()
    metrics = {
        "article": ARTICLE,
        "analysis_seed": ANALYSIS_SEED,
        "plot_seed": PLOT_SEED,
        "snapshot_date": SNAPSHOT_DATE,
        "storms_top_level_items": 17,
        "storms_expanded_recommendations": len(storms),
        "streams_recommendations": len(streams),
        "streams_manuscript_sections": streams["ManuscriptSection"].nunique(),
        "mimag_quality_levels": mimag["QualityLevel"].nunique(),
        "miuvig_mandatory_fields": len(miuvig_fields),
        "article44_mags": len(mag),
        "article44_high_quality": int(mag["Article44ExtendedTier"].eq("High quality").sum()),
        "article44_medium_quality": int(mag["Article44ExtendedTier"].eq("Medium quality").sum()),
        "article44_core_extended_agreement": int(mag["CoreVsExtendedAgreement"].sum()),
        "article54_uviqs": 46,
        "article54_miuvig_complete_fields": int(virus["Status"].eq("Complete").sum()),
        "article54_miuvig_partial_fields": int(virus["Status"].eq("Partial").sum()),
        "article54_miuvig_missing_fields": int(virus["Status"].eq("Missing").sum()),
        "readiness_status_counts": status_counts,
        "python": platform.python_version(),
        "source_records": len(sources["records"]),
    }
    (output / "analysis-metrics.json").write_text(
        json.dumps(metrics, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )
    methods = {
        "article": ARTICLE,
        "standards": {
            "STORMS": "v1.03; Zenodo 10.5281/zenodo.5714305",
            "STREAMS": "v1.0; Zenodo 10.5281/zenodo.15014818; STREAMS_final sheet",
            "MIMAG": "Bowers et al. 2017; DOI 10.1038/nbt.3893",
            "MIUViG": "Roux et al. 2019; DOI 10.1038/nbt.4306",
        },
        "boundary": (
            "STORMS and STREAMS guide study/manuscript reporting; MIMAG and MIUViG "
            "add entity-level records. Checklist completion is not a study-quality score."
        ),
        "streams_sheet_policy": "Only STREAMS_final was parsed; draft/development sheets were excluded.",
        "extended_mag_gate": "Article 44 adds GUNC pass to the MIMAG-derived tier gate.",
    }
    (output / "methods-contract.json").write_text(
        json.dumps(methods, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )
    (output / "data-NOTICE.txt").write_text(
        "STREAMS Figure 1 remains under publisher/rightsholder terms and is excluded "
        "from this repository's CC BY/MIT grant. Checklist workbooks and article XML "
        "retain their source terms. Derived tables and original diagrams are repository "
        "content. Article 44 MAGs and Article 54 virus fixtures are independent examples "
        "and cannot be combined into one biological claim.\n",
        encoding="utf-8",
    )
    print(f"prepared\t{output}\t{len(checklists)} checklist rows; {len(mag)} MAGs")


if __name__ == "__main__":
    main()
