#!/usr/bin/env python3
"""Prepare checksum-gated public inputs for Article 55."""

from __future__ import annotations

import argparse
import csv
import shutil
import time
import urllib.request
from pathlib import Path

from lxml import etree
from openpyxl import load_workbook

from article41_44_utils import dump_json, sha256, write_tsv


ASSETS = (
    {
        "AssetID": "cook-article-xml",
        "File": "PMC10926689.xml",
        "URL": "https://www.ebi.ac.uk/europepmc/webservices/rest/PMC10926689/fullTextXML",
        "Bytes": 140_775,
        "SHA256": "74d28ef0c7d21a14eaf80a8f517f5f9eadf5385f586b8f20cf280535491d4c4e",
        "DOI": "10.1099/mgen.0.001198",
        "License": "CC BY 4.0",
        "Role": "mock virome design, mapping policy, and interpretation",
    },
    {
        "AssetID": "deeppl-article-xml",
        "File": "PMC11521287.xml",
        "URL": "https://www.ebi.ac.uk/europepmc/webservices/rest/PMC11521287/fullTextXML",
        "Bytes": 157_363,
        "SHA256": "a28b8f2478ace607d497eb61f23e5c097107e30810d2ee077473cfd36b6f44eb",
        "DOI": "10.1371/journal.pcbi.1012525",
        "License": "CC BY 4.0",
        "Role": "predicted and experimentally supported phage lifecycle labels",
    },
    {
        "AssetID": "cook-supplement-workbook",
        "File": "mgen-10-1198-s002.xlsx",
        "URL": "https://www.ebi.ac.uk/biostudies/files/S-EPMC10926689/mgen-10-1198-s002.xlsx",
        "Bytes": 376_869,
        "SHA256": "4f7b44c6c276a6b2b60142543c5f2515917370ddac340a40bd85b37f13475642",
        "DOI": "10.1099/mgen.0.001198",
        "License": "CC BY 4.0",
        "Role": "15-phage input, library, coverage, depth, and diversity tables",
    },
    {
        "AssetID": "ena-run-filereport",
        "File": "ena-prjeb56639-filereport.tsv",
        "URL": (
            "https://www.ebi.ac.uk/ena/portal/api/filereport?"
            "accession=PRJEB56639&result=read_run&fields="
            "run_accession,sample_accession,experiment_accession,"
            "instrument_platform,instrument_model,library_name,library_strategy,"
            "library_source,library_layout,fastq_ftp,fastq_md5,fastq_bytes&format=tsv"
        ),
        "Bytes": 3_158,
        "SHA256": "c56c6b024f43dccbb56b22b8aeccdd7b3345a00081b7b19afce713b099eeb096",
        "DOI": "10.1099/mgen.0.001198",
        "License": "ENA public archive",
        "Role": "run accessions plus FASTQ URLs, publisher MD5, and byte counts",
    },
)


def retrieve(url: str, target: Path) -> None:
    target.parent.mkdir(parents=True, exist_ok=True)
    partial = target.with_suffix(target.suffix + ".part")
    for attempt in range(1, 6):
        try:
            request = urllib.request.Request(
                url, headers={"User-Agent": "metagenomics-best-practices/article55"}
            )
            with urllib.request.urlopen(request, timeout=180) as response:
                with partial.open("wb") as handle:
                    shutil.copyfileobj(response, handle, length=1024 * 1024)
            partial.replace(target)
            return
        except Exception:
            partial.unlink(missing_ok=True)
            if attempt == 5:
                raise
            time.sleep(2**attempt)


def load_manifest(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def read_one_fasta(path: Path) -> str:
    records = 0
    chunks: list[str] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        if line.startswith(">"):
            records += 1
        elif line.strip():
            chunks.append(line.strip().upper())
    if records != 1 or not chunks:
        raise ValueError(f"Expected exactly one FASTA record in {path}")
    return "".join(chunks)


def worksheet_rows(path: Path, sheet: str) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet not in workbook.sheetnames:
        raise KeyError(f"Missing worksheet {sheet}")
    rows = workbook[sheet].iter_rows(values_only=True)
    header = [str(value).strip() if value is not None else "" for value in next(rows)]
    output = []
    for values in rows:
        if not any(value is not None for value in values):
            continue
        output.append(dict(zip(header, values)))
    return output


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--project-root", type=Path, required=True)
    parser.add_argument("--work-dir", type=Path, required=True)
    args = parser.parse_args()
    root = args.project_root.resolve()
    work = args.work_dir.resolve()
    raw = root / "data/raw/article55"
    inputs = work / "input"
    inputs.mkdir(parents=True, exist_ok=True)

    asset_audit = []
    asset_manifest = []
    for asset in ASSETS:
        source = raw / str(asset["File"])
        if not source.is_file():
            retrieve(str(asset["URL"]), source)
        observed_bytes = source.stat().st_size
        observed_hash = sha256(source)
        passed = (
            observed_bytes == int(asset["Bytes"])
            and observed_hash == str(asset["SHA256"])
        )
        asset_audit.append(
            {
                "AssetID": asset["AssetID"],
                "File": asset["File"],
                "ExpectedBytes": asset["Bytes"],
                "ObservedBytes": observed_bytes,
                "ExpectedSHA256": asset["SHA256"],
                "ObservedSHA256": observed_hash,
                "ChecksumPass": passed,
            }
        )
        if not passed:
            raise RuntimeError(f"Checksum gate failed for {source}")
        shutil.copy2(source, inputs / str(asset["File"]))
        asset_manifest.append({**asset, "SourceIdentity": asset["AssetID"]})
    write_tsv(work / "asset-check-audit.tsv", asset_audit)
    write_tsv(work / "asset-manifest.tsv", asset_manifest)

    reference_manifest_path = root / "data/small/55-phage-reference-manifest.tsv"
    references = load_manifest(reference_manifest_path)
    if len(references) != 15 or len({row["phage_id"] for row in references}) != 15:
        raise RuntimeError("Article 55 reference manifest must contain 15 unique phages")
    reference_audit = []
    fasta_lines = []
    for row in references:
        accession = row["accession"]
        source = raw / "reference_genomes" / f"{accession}.fna"
        if not source.is_file():
            retrieve(row["download_url"], source)
        observed_hash = sha256(source)
        sequence = read_one_fasta(source)
        passed = (
            source.stat().st_size == int(row["fasta_bytes"])
            and observed_hash == row["fasta_sha256"]
            and len(sequence) == int(row["reference_sequence_length"])
        )
        reference_audit.append(
            {
                "PhageID": row["phage_id"],
                "Accession": accession,
                "PaperLengthBp": row["paper_genome_length"],
                "ReferenceLengthBp": len(sequence),
                "LengthDeltaBp": len(sequence) - int(row["paper_genome_length"]),
                "ObservedBytes": source.stat().st_size,
                "ObservedSHA256": observed_hash,
                "ChecksumAndLengthPass": passed,
            }
        )
        if not passed:
            raise RuntimeError(f"Reference gate failed for {source}")
        fasta_lines.append(f">{row['phage_id']}|{accession}")
        fasta_lines.extend(
            sequence[index : index + 80] for index in range(0, len(sequence), 80)
        )
    write_tsv(work / "reference-audit.tsv", reference_audit)
    (inputs / "cook15-phage-reference.fna").write_text(
        "\n".join(fasta_lines) + "\n", encoding="utf-8"
    )

    workbook = inputs / "mgen-10-1198-s002.xlsx"
    phage_rows = worksheet_rows(workbook, "Table_S1_Phages")[:15]
    coverage_rows = worksheet_rows(workbook, "Table_S4_Coverage_and_Depth")
    library_rows = worksheet_rows(workbook, "Table_S3_Libraries")
    if len(phage_rows) != 15 or len(coverage_rows) != 180:
        raise RuntimeError("Unexpected Cook supplementary workbook coordinate")

    by_id = {row["phage_id"]: row for row in references}
    proper_to_id = {
        "Escherichia phage vB_Eco_SLUR29": "vB_Eco_SLUR29",
        "PARMAL1": "PARMAL1",
        "Escherichia phage vB_Eco_mar002J2": "vB_Eco_mar002J2",
        "Bacteriophage HP1": "HP1",
        "Vibriophage vB_Vpa_sm033": "vB_Vpa_sm033",
        "Escherichia phage vB_Eco_mar003J3": "vB_Eco_mar003J3",
        "Escherichia phage vB_EcoS_swan01": "vB_EcoS_swan01",
        "KUW1": "KUW1",
        "Bacteriophage DSS3_PM1": "DSS3_PM1",
        "Escherichia phage vB_Eco_mar001J1": "vB_Eco_mar001J1",
        "Escherichia phage vB_Eco_mar005P1": "vB_Eco_mar005P1",
        "Vibriophage vB_VpaS_sm032": "vB_VpaS_sm032",
        "Coliphage phi-X174": "phix174",
        "Synechococcus phage S-RSM4": "S-RSM4",
        "Clostridium phage CDMH1": "CDMH1",
    }
    metadata = []
    for source in phage_rows:
        proper = str(source["Phage_proper_name"]).strip()
        phage_id = proper_to_id[proper]
        ref = by_id[phage_id]
        metadata.append(
            {
                "PhageID": phage_id,
                "Accession": ref["accession"],
                "ProperName": proper,
                "GenomeLengthPaperBp": int(source["Genome_length"]),
                "GenomeLengthReferenceBp": int(ref["reference_sequence_length"]),
                "GCPct": round(float(source["GC Content (%)"]), 6),
                "InputGenomeCopies": float(source["Number of  genome copies"]),
                "InputTPM": float(source["TPM*"]),
                "DeepPLPrediction": ref["deeppl_prediction"],
                "PhaTYPPrediction": ref["phatyp_prediction"],
                "ConfirmedLifecycle": ref["confirmed_lifecycle"],
                "LifecycleSourceNote": ref["lifecycle_source_note"],
            }
        )
    write_tsv(work / "phage-metadata.tsv", metadata)

    published = []
    for row in coverage_rows:
        published.append(
            {
                "Dataset": row["Dataset"],
                "Library": row["Library"],
                "Platform": row["Read_type"],
                "PhageID": row["Phage"],
                "BreadthPct": row["Coverage"],
                "MeanDepthX": row["Depth"],
                "InputGenomeCopies": row["Genome_Copies"],
                "InputCPM": row["CPM"],
            }
        )
    write_tsv(work / "published-coverage-depth.tsv", published)

    ena_path = inputs / "ena-prjeb56639-filereport.tsv"
    with ena_path.open(encoding="utf-8", newline="") as handle:
        ena_rows = list(csv.DictReader(handle, delimiter="\t"))
    illumina_runs = [row for row in ena_rows if row["instrument_platform"] == "ILLUMINA"]
    expected_runs = {"ERR10359653", "ERR10359656", "ERR10359658"}
    if {row["run_accession"] for row in illumina_runs} != expected_runs:
        raise RuntimeError("Unexpected PRJEB56639 Illumina run coordinate")
    write_tsv(work / "ena-illumina-run-manifest.tsv", illumina_runs)
    illumina_libraries = [
        {
            "Dataset": f"lib{int(row['New_Library_Number'])}_illumina",
            "Library": int(row["New_Library_Number"]),
            "ReadPairs": int(row["Number_Reads"]),
            "BasePairs": int(row["Base_Pairs"]),
            "RunAccession": row["ENA_Run_Accession"],
            "MDA": "No",
        }
        for row in library_rows
        if row["Sequencing_Platform"] == "Illumina"
        and str(row["New_Library_Number"]) != "Pooled"
    ]
    write_tsv(work / "illumina-library-metadata.tsv", illumina_libraries)

    cook_text = " ".join(
        " ".join(
            etree.parse(str(inputs / "PMC10926689.xml")).getroot().itertext()
        ).split()
    )
    assertions = [
        ("BBMap identity and ambiguity", "Bbmap v38.69 at 90 % minimum ID and the ambiguous=all flag" in cook_text),
        ("primary presence gate", "≥ 1×coverage across ≥ 75 % of contig length" in cook_text),
        ("later 70 percent wording", "detected by read mapping across ≥70 % of its length" in cook_text),
        ("four universal nondetections", "CDMH1, HP1, vB_Eco_mar005P1 and ΦX174" in cook_text),
        ("Illumina no MDA", "For Illumina sequencing, no MDA was used" in cook_text),
    ]
    write_tsv(
        work / "author-source-assertions.tsv",
        [
            {"Assertion": label, "Pass": passed, "Source": "PMC10926689"}
            for label, passed in assertions
        ],
    )
    if not all(passed for _, passed in assertions):
        raise RuntimeError("One or more Cook source assertions failed")

    deeppl_text = " ".join(
        " ".join(
            etree.parse(str(inputs / "PMC11521287.xml")).getroot().itertext()
        ).split()
    )
    lifecycle_assertions = [
        ("HP1 temperate", "HP1 NC_001697 32,355 Lysogenic Lysogenic Temperate" in deeppl_text),
        ("J1 disagreement", "J1 LR027388 50,343 Lysogenic Lytic Temperate" in deeppl_text),
        ("J2 disagreement", "J2 LR027385 50,343 Lysogenic Lytic Temperate" in deeppl_text),
        ("SWAN disagreement", "SWAN LT841304 50,865 Lysogenic Lytic Temperate" in deeppl_text),
        ("unknown labels retained", "KUW1 OQ376857 44,509 Lytic Lytic NA NA" in deeppl_text),
    ]
    write_tsv(
        work / "lifecycle-source-assertions.tsv",
        [
            {"Assertion": label, "Pass": passed, "Source": "PMC11521287 Table 3"}
            for label, passed in lifecycle_assertions
        ],
    )
    if not all(passed for _, passed in lifecycle_assertions):
        raise RuntimeError("One or more DeepPL source assertions failed")

    write_tsv(
        work / "input-lineage.tsv",
        [
            {
                "Input": "mgen-10-1198-s002.xlsx",
                "Coordinate": "Tables S1, S3, S4",
                "Origin": "Cook et al. 2024 supplement",
                "NextStep": "published mapping audit and threshold sensitivity",
            },
            {
                "Input": "15 accession-locked FASTA files",
                "Coordinate": "55-phage-reference-manifest.tsv",
                "Origin": "NCBI Nucleotide",
                "NextStep": "geNomad v1.9 taxonomy and 95/85 vOTU clustering",
            },
            {
                "Input": "PMC11521287.xml",
                "Coordinate": "Table 3",
                "Origin": "DeepPL paper",
                "NextStep": "prediction-versus-confirmed lifecycle ledger",
            },
        ],
    )
    dump_json(
        work / "run-contract.json",
        {
            "article": 55,
            "seed": 20260755,
            "primary_platform": "Illumina without MDA",
            "primary_presence_min_depth": 1,
            "primary_presence_min_breadth_pct": 75,
            "author_mapping_min_identity_pct": 90,
            "author_multimapping_policy": "ambiguous=all",
            "votu_min_ani_pct": 95,
            "votu_min_alignment_fraction_shorter_pct": 85,
            "random_output_requested": False,
        },
    )
    print(f"Article 55 inputs verified in {work}")


if __name__ == "__main__":
    main()
